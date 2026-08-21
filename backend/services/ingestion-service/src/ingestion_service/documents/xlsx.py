from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path

from openpyxl import load_workbook

from ingestion_service.domain import BlockType, DocumentUnit, ExtractedBlock


class XLSXReader:
    def supports(self, path: Path, content_type: str) -> bool:
        return path.suffix.lower() == ".xlsx" or content_type == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    def units(self, path: Path) -> Iterator[DocumentUnit]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for number, sheet in enumerate(workbook.worksheets, start=1):
                rows: list[str] = []
                for row in sheet.iter_rows(values_only=True):
                    values = [str(value).strip() if value is not None else "" for value in row]
                    if any(values):
                        rows.append(" | ".join(values))
                yield DocumentUnit(
                    number=number,
                    label=sheet.title,
                    blocks=[
                        ExtractedBlock(
                            block_type=BlockType.TABLE,
                            text="\n".join(rows),
                            section_path=[sheet.title],
                        )
                    ],
                    metadata={"format": "xlsx", "sheet": sheet.title},
                )
        finally:
            workbook.close()

